import openpyxl as op
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import csv

# For testing
xlsx_filename = "test.xlsx"

def check_for_overview_sheet(filename):
    """
    Checks if the given xlsx file contains a sheet with a certain name.
    Make changes to the function to adapt it to your own usage.
    The file needs to be within the same folder as the script.
    
    Arguments:
        filename: The name of the file within the same directory as the script.
    
    Returns:
        True if the specified sheet exists, False otherwise.
    """
    try:
        doc = op.load_workbook(filename)
        return "Oversikt" in doc.sheetnames

    except Exception as e:
        print(f"Error loading document: {e}")
        return False

def create_new_doc(file_name):
    """
    Creates a new xlsx document with all the necessary sheets and data.
    
    Arguments:
        name: Name of the document
    
    Returns:
        Nothing - A helper function to set up a new document
    """
    #Remove any extensions
    basename = file_name.split(".")[0]
    file_name_ext = f"{basename}.xlsx"
    # List of all the sheets except Overview which will be handled separately
    sheets = [ "Inntekter", "Sparing", "Fond", "PC Relatert", "Elektronikk", 
              "Spill", "Klær", "Kjøretøy", "Prosjekter", "Husholdning", "TakeAway", "Mat", 
              "Art", "Annet"]
    
    workbook = op.Workbook()

    # Renaming of the default sheet page
    overview_sheet = workbook.active
    overview_sheet.title = "Oversikt"

    # Define named styles for the overview page
    center_aligned = Alignment(horizontal="center", vertical="center")
    total_font = Font(name="Calibri", size=10)
    title_style = NamedStyle(name="title_style", font=Font(name="Calibri", size=20, bold=True))
    subtitle_style = NamedStyle(name="subtitle_style", font=Font(name="Calibri", size=12, bold=True, italic=True))
    sheet_name_style = NamedStyle(name="sheet_name_style", font=Font(name="Calibri", size=11, bold=True, italic=True))
    total_positive = NamedStyle(name="total_positive", font=total_font, fill=PatternFill(start_color="00a933", end_color="00a933", fill_type="solid"))
    total_fond = NamedStyle(name="total_fond", font=total_font, fill=PatternFill(start_color="2a6099", end_color="2a6099", fill_type="solid"))
    total_negative = NamedStyle(name="total_negative", font=total_font, fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"))

    # Adding the named styles to the workbook
    if "title_style" not in workbook.named_styles:
        workbook.add_named_style(title_style)
    if "subtitle_style" not in workbook.named_styles:
        workbook.add_named_style(subtitle_style)
    if "sheet_name_style" not in workbook.named_styles:
        workbook.add_named_style(sheet_name_style)
    if "total_positive" not in workbook.named_styles:
        workbook.add_named_style(total_positive)
    if "total_fond" not in workbook.named_styles:
        workbook.add_named_style(total_fond)
    if "total_negative" not in workbook.named_styles:
        workbook.add_named_style(total_negative)

    # Title
    overview_sheet["A1"] = "Oversikt"
    overview_sheet["A1"].style = "title_style"
    overview_sheet["A1"].alignment = center_aligned
    overview_sheet.row_dimensions[1].height = 29
    overview_sheet.column_dimensions["A"].width = 24

    # Subtitles
    overview_sheet["A4"] = "Kategori:"
    overview_sheet["A4"].style = "subtitle_style"
    overview_sheet["A4"].alignment = center_aligned
    overview_sheet["B4"] = "Total:"
    overview_sheet["B4"].style = "subtitle_style"
    overview_sheet["B4"].alignment = center_aligned

    # Category/sheet texts
    next_row = 5

    #loop through the sheets list to cut down on manual typing
    for name in sheets:
        a_cell = f"A{next_row}"
        overview_sheet[a_cell] = name
        overview_sheet[a_cell].style = "sheet_name_style"
        overview_sheet[a_cell].alignment = center_aligned

        # set the formula to sum up for the total
        b_cell = f"B{next_row}"
        overview_sheet[b_cell] = f"=SUM('{name}'!B:B)"
        overview_sheet[b_cell].number_format = "# ##0,00"

        # Styling for the total cells
        if name == "Inntekter":
            overview_sheet[b_cell].style = "total_positive"
        elif name == "Sparing":
            overview_sheet[b_cell].style = "total_positive"
        elif name == "Fond":
            overview_sheet[b_cell].style = "total_fond"
        else:
            overview_sheet[b_cell].style = "total_negative"

        overview_sheet[b_cell].alignment = center_aligned
        
        next_row += 1
    
    # Create the specified sheets
    for name in sheets:
        sheet = workbook.create_sheet(title=name)
        sheet["A1"] = "Dato:"
        sheet["B1"] = "Beløp:"
        sheet["C1"] = "Beskrivelse:"
        sheet.column_dimensions["A"].width = 14
        sheet.column_dimensions["B"].width = 14
        for cell in sheet["A"]:
            cell.number_format = "DD.MM"
        for cell in sheet["B"]:
            cell.number_format = "# ##0,00"

    # Save the document
    print("Saving the document as ", file_name_ext)
    workbook.save(file_name_ext)

    return
    

def filter_csv(filename):
    """
    Opens up a CSV file and stores the columns specified by the code.
    Tweak the code to your own liking. 
    
    Arguments:
        Filename: The name of the file within the same directory as the script.
    
    Returns:
        A list of dictionary objects containing the information within the CSV.
    """
    filtered_transactions = []

    with open(filename, newline="",) as f:
        reader = csv.DictReader(f, delimiter=";")
        for row in reader:
            # Check if there is a value in "Beløp ut" for converting to float
            if row["Beløp ut"]:
                # Remove the - from the value and convert to a float
                amount = float(row["Beløp ut"].replace("-", ""))
            # Check if there is a value in "Beløp inn" for converting to float
            elif row["Beløp inn"]:
                amount = float(row["Beløp inn"])
            else:
                #If there is no Amount out or in, then there is no more lines to process.
                #Cuts out the extra information in my particular CSV
                break

            #Convert the date into a datetime obj
            try:
                date_obj = datetime.strptime(row["Utført dato"], "%d.%m.%Y")
            except ValueError:
                print(f"Invalid date format: {row['Utført dato']}")
                continue

            # Write the transaction to a dictionary object to store in the list
            transaction = {
                "Dato": date_obj,
                "Beskrivelse": row["Beskrivelse"],
                "Beløp": amount
            }

            filtered_transactions.append(transaction)
    sorted_transactions = sorted(filtered_transactions, key=lambda x: x["Dato"])    
    return sorted_transactions

def get_excel_categories(filename):
    """
    Takes in an excel file and fetches the names of all the sheets (except the overview sheet).
    
    Arguments:
        filename: filepath to an excel file
        
    Returns:
        List of strings containing the names of the different sheets
    """
    # First entry of the category dropdown menu should be a "Choose category" option
    categories = ["Velg Kategori"]
    workbook = op.load_workbook(filename=filename, read_only=True)

    # Loop over the different sheets in the excel file - Exclude the overview sheet.
    for sheet in workbook.sheetnames:
        if sheet != "Oversikt": 
            categories.append(sheet)

    return categories

def add_new_category(filename, category):
    """
    Takes in an excel file and category to be able to add a new sheet to an existing excel document.
    
    Arguments:
        filename: filepath to an excel file.
        category: string containing a name for a new category
        
    Returns:
        Nothing
    """
    workbook = op.load_workbook(filename=filename)
    # Convert all the sheet names into lower case for comparison
    sheetnames_lower = [sheet.lower() for sheet in workbook.sheetnames]

    if category.lower() not in sheetnames_lower:
        new_sheet = workbook.create_sheet(title=category)
        new_sheet["A1"] = "Dato:"
        new_sheet["B1"] = "Beløp:"
        new_sheet["C1"] = "Beskrivelse:"
        overview_sheet = workbook["Oversikt"]
        next_row = overview_sheet.max_row + 1
        overview_sheet[f"A{next_row}"] = category
        overview_sheet[f"A{next_row}"].style = "sheet_name_style"
        overview_sheet[f"A{next_row}"].alignment = Alignment(horizontal="center", vertical="center")

        overview_sheet[f"B{next_row}"] = f"=SUM('{category}'!B:B)"
        overview_sheet[f"B{next_row}"].style = "total_negative"
        overview_sheet[f"B{next_row}"].alignment = Alignment(horizontal="center", vertical="center")

        workbook.save(filename)

    return

def process_transactions(transactions):
    """
    Test function:
        Takes in a list of transactions and let's the user add a category and comment to the transaction.
    
    Arguments:
        transactions: Should be a list of dictionaries returned from the function 'filter_csv'
    
    Returns:
        A new list of dictionaries with Date, Amount, Comment and Category
    """
    #For now, add it to "Annet" on category and "Test" for comment to see if it works
    new_transactions = []
    for entry in transactions:
        #formatted_date = f"{entry["Dato"].day}.{entry["Dato"].month}"
        new_entry = {
            "Dato": entry["Dato"],
            "Beløp": entry["Beløp"],
            "Beskrivelse": "Test",
            "Kategori": "Annet"
        }
        new_transactions.append(new_entry)
    return new_transactions

def save_document(transactions, filename):
    """
    Takes in a list of transactions with Date, Amount, Comment and Category to append to a xlsx file
    
    Arguments:
        transactions: Needs to be a list of dictionaries with Date, Amount, Comment and Category - Use process_transactions()
        
    Returns:
        Nothing - The function will try to append the transactions into the document and try to save it.
    """
    try:
        workbook = op.load_workbook(filename)
    except FileNotFoundError:
        print("File not found")
        return
    
    # Check if there is a style named "date_style" in the file already
    if "date_style" not in workbook.named_styles:
        # Add the style to get the desired format and font for dates
        date_style = NamedStyle(name="date_style", number_format="DD.MM.YYYY")
        date_style.font = Font(name="Calibri")
        workbook.add_named_style(date_style)

    for entry in transactions:
        category = entry["Kategori"]
        try:
            #Get data from the corresponding category sheet.
            worksheet = workbook[category]
        except KeyError:
            worksheet = workbook.create_sheet(title=category)

        #Find the next available row to append data
        next_row = worksheet.max_row + 1
        
        date_obj = datetime.strptime(entry["Dato"], "%Y-%m-%d %H:%M:%S")
        date_cell = worksheet.cell(row=next_row, column=1, value=date_obj)
        date_cell.style = "date_style"
        number_cell = worksheet.cell(row=next_row, column=2, value=entry["Beløp"])
        number_cell.number_format = "# ###,00"          # Unfortunately doesn't work as I expect it to
        worksheet.cell(row=next_row, column=3, value=entry["Beskrivelse"])
    
    workbook.save(filename)

# Test to see if it works
#csv_filename = "transaksjoner_test.csv"
#filtrert = filter_csv(csv_filename)
#klargjort = process_transactions(filtrert)
#save_document(klargjort)
#create_new_doc("new_test")
#create_new_doc("example")
#create_new_doc("example.xlsx")
#create_new_doc("example.test")
#create_new_doc("example.test.xlsx")